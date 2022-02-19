# -*- coding: UTF-8 -*-
import openpyxl
import requests
from requests.exceptions import ReadTimeout, ConnectionError

def get_detail(uid, sa, wa, i):
    headers = {
        'authority': 'weibo.com',
        'pragma': 'no-cache',
        'cache-control': 'no-cache',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="96", "Google Chrome";v="96"',
        'x-xsrf-token': 'RAe1GEyd4RK0b_Ro7gbXEdAB',
        'traceparent': '00-477fdf7e5a75c0b883d5a2feadfb2900-e2c2d71e4b24e1c5-00',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36',
        'accept': 'application/json, text/plain, */*',
        'x-requested-with': 'XMLHttpRequest',
        'sec-ch-ua-platform': '"macOS"',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://weibo.com/dazzlefashion?is_hot=1',
        'accept-language': 'zh-CN,zh;q=0.9',
        # 填写您的cookie
        'cookie': '',
    }

    params = (
        ('custom', str(uid)),
    )
    while True:
        try:
            response = requests.get('https://weibo.com/ajax/profile/info', headers=headers, params=params)
            break
        except ReadTimeout as e:
            print(e)
            continue
        except ConnectionError as e:
            print(e)
            continue
    data = response.json()
    detail = data.get('data', {}).get('user', {})
    if not detail:
        return
    gender = ''
    if detail.get('gender', '') == 'f':
        gender = '女'
    elif detail.get('gender', '') == 'm':
        gender = '男'
    content = {
        "user_name": detail.get('screen_name', ''),
        "uid": detail.get('idstr', ''),
        "gender": gender,
        'follow_count': detail.get('followers_count'),
        'friends_count': detail.get('friends_count'),
        'verified_reason': detail.get('verified_reason', ''),
        'description': detail.get('description', ''),
        'location': detail.get('location', ''),
        'statuses_count': detail.get('statuses_count')
    }
    print(content)
    sa.cell(row=i, column=1).value = content.get('user_name', '')
    sa.cell(row=i, column=2).value = content.get('uid', '')
    sa.cell(row=i, column=3).value = content.get('gender', '')
    sa.cell(row=i, column=4).value = content.get('follow_count', '')
    sa.cell(row=i, column=5).value = content.get('friends_count', '')
    sa.cell(row=i, column=6).value = content.get('verified_reason', '')
    sa.cell(row=i, column=7).value = content.get('description', '')
    sa.cell(row=i, column=8).value = content.get('location', '')
    sa.cell(row=i, column=9).value = content.get('statuses_count', '')
    wa.save('weibo.xlsx')


if __name__ == '__main__':
    wa = openpyxl.load_workbook('weibo.xlsx')
    sa = wa['weibo']
    sa.cell(row=1, column=1).value = '作者名字'
    sa.cell(row=1, column=2).value = 'user_id'
    sa.cell(row=1, column=3).value = '性别'
    sa.cell(row=1, column=4).value = '粉丝数'
    sa.cell(row=1, column=5).value = '关注数'
    sa.cell(row=1, column=6).value = '身份'
    sa.cell(row=1, column=7).value = '描述'
    sa.cell(row=1, column=8).value = '位置'
    sa.cell(row=1, column=9).value = '微博总数'
    # excel 表名
    wb = openpyxl.load_workbook('UserUid.xlsx')
    # sheet名
    sh = wb['Sheet1']
    for i in range(2726, 5464):
        uid = sh.cell(row=i, column=3).value
        if not uid:
            continue
        get_detail(uid, sa, wa, i)
    wb.close()
    wa.close()
